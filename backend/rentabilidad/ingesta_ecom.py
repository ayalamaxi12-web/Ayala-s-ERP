"""Adaptador de solo lectura sobre el Excel que hoy se descarga manualmente
del reporte de ECOM (por Fecha de Creación, rango personalizado — típicamente
del 23 al 22 de cada mes). Reemplaza la limpieza y validación manual descripta
por Maxx (2026-07-31) antes de que el Excel llegue al motor de Rentabilidad.

Mismo principio que `ingesta_tactica.py`: solo lectura, separa la lectura de
archivo (`leer_filas`, inyectable en tests) de la traducción a
`LineaEcomInput`. El motor (`RentabilidadEcomCalculator`) no se toca ni se
acopla al Excel — este adaptador es reemplazable por API o Chrome sin tocar
el cálculo (así lo pidió Maxx explícitamente).

Verificado contra archivos reales de Maxx (2026-07-31, no versionados en el
repo): la fórmula de `RentabilidadEcomCalculator` reproduce exactamente
Neto/Costo Total/Rentabilidad de su planilla procesada, centavo a centavo,
en las filas comparadas.

Reglas de este adaptador (autorizadas por Maxx, 2026-07-31):

1. **Limpieza del archivo**: el reporte trae 2 filas de encabezado
   informativo antes de la fila de encabezados reales (fila 3). Además, en
   la práctica el archivo real trae de más filas totalmente vacías después
   de los datos (padding del export) — se ignoran, no es una regla de
   negocio sino higiene de lectura.
2. **Exclusión por Estado de Pago**: solo `Cobrado`/`Cobro Parcial`
   participan (se calculan igual entre sí, RENTABILIDAD_FUNCIONAL.md §10
   v2.1). Cualquier otro valor (`Reembolsado`, `Sin cobro`, `En Mediación`,
   o cualquier estado no reconocido) se excluye completamente — lista
   blanca, no lista negra, para no arriesgar incluir por default un estado
   nuevo que en realidad signifique "todavía no impacta". Si el período se
   recalcula completo, una orden que cambió de estado entra sola.
3. **Validación de costo**: si `Costo Sin Iva` es cero o no viene, NUNCA se
   asume costo 0 — se marca la línea como incidencia
   (`COSTO_NO_RESUELTO`) para revisión manual y no se calcula. El caso
   típico (SKU madre en vez de la variante vendida) requiere el detalle de
   la venta, que este adaptador no tiene — no se inventa una resolución.
4. **Postventa** (RMA): `Precio Final` y `Precio Sin IVA` se fuerzan a 0,
   sin importar lo que traiga el Excel — la rentabilidad de esas líneas
   refleja únicamente la pérdida del costo del producto.
5. **TC**: a diferencia de Táctica, el Excel de ECOM no trae una cotización
   propia por línea — Maxx aplica un único TC (BNA) a todo el período al
   momento de procesarlo (confirmado contra archivo real: mismo valor en
   todas las filas de un período). Por eso se recibe como parámetro de
   `lineas()`, no se lee de ninguna columna.

Todo lo demás (PM, categorías, responsables, listas, comparaciones,
reportes personales) queda deliberadamente fuera — Maxx fue explícito en
que este adaptador solo entrega líneas listas para el motor.
"""
from dataclasses import dataclass
from decimal import Decimal
from typing import Callable

from .calculators import LineaEcomInput

# §10 v2.1 RENTABILIDAD_FUNCIONAL.md — únicos dos estados que participan.
# Lista blanca a propósito (ver docstring del módulo, punto 2).
ESTADOS_PAGO_QUE_PARTICIPAN = {"Cobrado", "Cobro Parcial"}

CANAL_POSVENTA = "Posventa"

# Candidatos de título de columna — el Excel crudo de ECOM usa "Precio Neto"
# para lo que el diccionario de datos (RENTABILIDAD_FUNCIONAL.md §7.7) llama
# "Precio SIN IVA" (columna Q); se resuelve por título, no por letra/índice
# fijo (§2.2 RENTABILIDAD_IMPLEMENTACION.md), para no romper si ECOM
# renombra la columna en una futura exportación.
_COL_NUMERO_ORDEN = "Número Orden"
_COL_SKUS_VENDIDOS = "Sku's Vendidos"
_COL_ESTADO_PAGO = "Estado Pago"
_COL_CANAL_VENTA = "Canal De Venta"
_COL_COSTO_SIN_IVA = "Costo Sin Iva (total de productos)"
_COL_COMISION_VENTA = "Comisión Venta"
_COL_COSTO_ENVIO = "Costo Envío"
_COL_PRECIO_FINAL = "Precio Final"
_COLS_PRECIO_SIN_IVA = ("Precio SIN IVA", "Precio Neto")


@dataclass
class FilaEcom:
    """Una orden ya limpia — lista para `a_linea_input()`, salvo que tenga
    `incidencia` (no se calcula, se revisa a mano)."""

    numero_orden: str
    skus_vendidos: str
    canal_de_venta: str | None
    estado_pago: str
    costo_sin_iva: Decimal
    comision_venta: Decimal
    costo_envio: Decimal
    precio_sin_iva: Decimal
    precio_final: Decimal
    tc: Decimal
    incidencia: str | None = None

    def a_linea_input(self) -> LineaEcomInput:
        return LineaEcomInput(
            numero_orden=self.numero_orden,
            costo_sin_iva=self.costo_sin_iva,
            comision_venta=self.comision_venta,
            costo_envio=self.costo_envio,
            precio_sin_iva=self.precio_sin_iva,
            precio_final=self.precio_final,
            tc=self.tc,
        )


@dataclass
class ResultadoIngestaEcom:
    """Separa lo calculable de lo que necesita ojos humanos — nada se
    descarta en silencio (§1.2 RENTABILIDAD_IMPLEMENTACION.md: la
    exclusión debe ser auditable)."""

    lineas: list[FilaEcom]
    excluidas_por_estado_pago: list[FilaEcom]
    incidencias_costo: list[FilaEcom]


def _valor(fila: tuple, idx: int | None):
    return fila[idx] if idx is not None and idx < len(fila) else None


def _decimal(v) -> Decimal:
    if v in (None, ""):
        return Decimal(0)
    return Decimal(str(v))


def _texto(v) -> str:
    return str(v).strip() if v is not None else ""


def _mapa_columnas(encabezados: tuple) -> dict[str, int]:
    return {_texto(h): i for i, h in enumerate(encabezados) if h}


def _indice(mapa: dict[str, int], candidatos) -> int | None:
    for candidato in candidatos:
        if candidato in mapa:
            return mapa[candidato]
    return None


def leer_filas_excel(path: str) -> list[dict]:
    """Import perezoso de openpyxl — igual que `gsheets.get_client()` con
    gspread, para que el resto de `rentabilidad/` no dependa de este driver
    para correr sus tests.

    Descarta las 2 filas de encabezado informativo del reporte (fila 1:
    "Periodo" + rango de fechas; fila 2: vacía), usa la fila 3 como
    encabezados reales, e ignora las filas completamente vacías que trae el
    export después de los datos (`Número Orden` vacío)."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    filas = list(ws.iter_rows(values_only=True))
    mapa = _mapa_columnas(filas[2])
    idx_orden = _indice(mapa, (_COL_NUMERO_ORDEN,))

    resultado = []
    for fila in filas[3:]:
        if not _valor(fila, idx_orden):
            continue
        resultado.append({titulo: _valor(fila, i) for titulo, i in mapa.items()})
    return resultado


LeerFilas = Callable[[str], list[dict]]


def _fila_desde_row(row: dict, tc: Decimal) -> FilaEcom:
    estado_pago = _texto(row.get(_COL_ESTADO_PAGO))
    canal = _texto(row.get(_COL_CANAL_VENTA)) or None
    costo = _decimal(row.get(_COL_COSTO_SIN_IVA))

    if canal == CANAL_POSVENTA:
        # Regla de Maxx (2026-07-31): RMA — se fuerza a 0 sin importar el
        # dato de origen, la rentabilidad de la línea es pura pérdida de costo.
        precio_sin_iva = Decimal(0)
        precio_final = Decimal(0)
    else:
        precio_sin_iva = _decimal(next(
            (row.get(c) for c in _COLS_PRECIO_SIN_IVA if row.get(c) is not None), None
        ))
        precio_final = _decimal(row.get(_COL_PRECIO_FINAL))

    incidencia = "COSTO_NO_RESUELTO" if costo <= 0 else None

    orden_raw = row.get(_COL_NUMERO_ORDEN)
    numero_orden = str(int(orden_raw)) if isinstance(orden_raw, float) and orden_raw.is_integer() else str(orden_raw)

    return FilaEcom(
        numero_orden=numero_orden,
        skus_vendidos=_texto(row.get(_COL_SKUS_VENDIDOS)),
        canal_de_venta=canal,
        estado_pago=estado_pago,
        costo_sin_iva=costo,
        comision_venta=_decimal(row.get(_COL_COMISION_VENTA)),
        costo_envio=_decimal(row.get(_COL_COSTO_ENVIO)),
        precio_sin_iva=precio_sin_iva,
        precio_final=precio_final,
        tc=tc,
        incidencia=incidencia,
    )


class EcomExcelAdapter:
    """Solo lectura. `tc` se recibe como parámetro (ver punto 5 del
    docstring del módulo) — no se lee de ninguna columna del Excel."""

    def __init__(self, leer_filas: LeerFilas | None = None):
        self._leer_filas = leer_filas or leer_filas_excel

    def procesar(self, path: str, tc: Decimal) -> ResultadoIngestaEcom:
        lineas: list[FilaEcom] = []
        excluidas: list[FilaEcom] = []
        incidencias: list[FilaEcom] = []

        for row in self._leer_filas(path):
            fila = _fila_desde_row(row, tc)
            if fila.estado_pago not in ESTADOS_PAGO_QUE_PARTICIPAN:
                excluidas.append(fila)
            elif fila.incidencia is not None:
                incidencias.append(fila)
            else:
                lineas.append(fila)

        return ResultadoIngestaEcom(
            lineas=lineas, excluidas_por_estado_pago=excluidas, incidencias_costo=incidencias,
        )
